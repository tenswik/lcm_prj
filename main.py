import math
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
from math import sqrt, sin, cos, tan, radians, degrees, atan
from datetime import datetime


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.lib.colors import black, gray, darkblue, white, lightgrey
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ConveyorCalculator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Программа для тягового расчета ленточного конвейера")

        # === АДАПТАЦИЯ РАЗМЕРОВ ОКНА ===
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Устанавливаем размеры окна (60% от экрана, максимум 700x600)
        window_width = min(int(screen_width * 0.60), 700)
        window_height = min(int(screen_height * 0.70), 600)

        # Минимальные размеры
        min_width = 550
        min_height = 450

        self.root.geometry(f"{window_width}x{window_height}")
        self.root.minsize(min_width, min_height)

        # Центрирование окна на экране
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"+{x}+{y}")

        self.root.resizable(True, True)

        # Загрузка настроек, истории и счётчика расчётов
        self.load_settings()
        self.load_history()
        self.load_calculation_counter()

        # Регистрация кириллического шрифта для PDF
        self.register_cyrillic_font()

        # Создание основных фреймов
        self.create_widgets()

    def register_cyrillic_font(self):
        """Регистрация шрифта с поддержкой кириллицы для PDF"""
        if PDF_AVAILABLE:
            try:
                font_paths = [
                    'DejaVuSans.ttf',
                    'C:/Windows/Fonts/arial.ttf',
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                    '/System/Library/Fonts/Arial.ttf'
                ]

                for font_path in font_paths:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                        self.cyrillic_font = 'CyrillicFont'
                        return

                self.cyrillic_font = 'Helvetica'
            except:
                self.cyrillic_font = 'Helvetica'
        else:
            self.cyrillic_font = 'Helvetica'

    def load_settings(self):
        """Загрузка сохраненных настроек"""
        try:
            with open('settings.json', 'r', encoding='utf-8') as f:
                self.settings = json.load(f)
        except FileNotFoundError:
            self.settings = {
                "угол_наклона_бокового_ролика": 20,
                "расстояние_между_роликом_и_опорой": 1.2,
                "количество_роликов": 10,
                "масса_одного_ролика": 15.0
            }

    def save_settings(self):
        """Сохранение настроек"""
        with open('settings.json', 'w', encoding='utf-8') as f:
            json.dump(self.settings, f, ensure_ascii=False, indent=2)

    def load_history(self):
        """Загрузка истории расчётов"""
        try:
            with open('calculations_history.json', 'r', encoding='utf-8') as f:
                self.history = json.load(f)
        except FileNotFoundError:
            self.history = []

    def save_history(self):
        """Сохранение истории расчётов"""
        with open('calculations_history.json', 'w', encoding='utf-8') as f:
            json.dump(self.history, f, ensure_ascii=False, indent=2)

    def load_calculation_counter(self):
        """Загрузка счётчика расчётов"""
        try:
            with open('calculation_counter.json', 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.calculation_number = data.get('counter', 0)
        except FileNotFoundError:
            self.calculation_number = 0

    def save_calculation_counter(self):
        """Сохранение счётчика расчётов"""
        with open('calculation_counter.json', 'w', encoding='utf-8') as f:
            json.dump({'counter': self.calculation_number}, f, ensure_ascii=False, indent=2)

    def increment_calculation_counter(self):
        """Увеличение счётчика расчётов на 1"""
        self.calculation_number += 1
        self.save_calculation_counter()
        return self.calculation_number

    def add_to_history(self, data, results, calc_number):
        """Добавление расчёта в историю"""
        calculation_record = {
            "calculation_number": calc_number,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "input_data": data,
            "results": {
                "operating_conditions": self.get_operating_conditions(data),
                "geometry": {
                    "length": data.get('length', 0),
                    "angle": data.get('angle', 0),
                    "height": data.get('height', 0),
                    "distance": data.get('distance', 0)
                },
                "min_belt_width": self.get_min_belt_width(data),
                "capacity": self.get_calculated_capacity(data),
                "cross_section_area": self.get_cross_section_area(data),
                "weather_loads": self.get_weather_loads(data),
                "resistance_force": self.get_resistance_force(data),
                "drive_power": self.get_drive_power(data)
            }
        }

        self.history.append(calculation_record)
        self.save_history()

    def create_widgets(self):
        """Создание виджетов интерфейса"""
        main_container = tk.Frame(self.root)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.canvas = tk.Canvas(main_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw", width=640)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        notebook = ttk.Notebook(self.scrollable_frame)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        main_frame = ttk.Frame(notebook)
        notebook.add(main_frame, text="Основные данные")

        settings_frame = ttk.Frame(notebook)
        notebook.add(settings_frame, text="Настройки")

        history_frame = ttk.Frame(notebook)
        notebook.add(history_frame, text="История расчётов")

        self.fill_main_frame(main_frame)
        self.fill_settings_frame(settings_frame)
        self.fill_history_frame(history_frame)
        button_frame = tk.Frame(self.root, bg="#f0f0f0", pady=5)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)

        calc_button = tk.Button(button_frame, text="Выполнить расчет", command=self.calculate,
                                bg="#4CAF50", fg="white", font=("Arial", 10, "bold"),
                                padx=15, pady=5)
        calc_button.pack(side=tk.LEFT, padx=5)

        history_button = tk.Button(button_frame, text="Обновить историю", command=self.refresh_history,
                                   padx=10, pady=5)
        history_button.pack(side=tk.LEFT, padx=5)

        clear_history_button = tk.Button(button_frame, text="Очистить историю", command=self.clear_history,
                                         padx=10, pady=5)
        clear_history_button.pack(side=tk.LEFT, padx=5)

        self.counter_label = tk.Label(self.root, text=f"Следующий расчёт №: {self.calculation_number + 1}",
                                      font=("Arial", 10, "bold"), bg="#f0f0f0", pady=3)
        self.counter_label.pack(side=tk.BOTTOM, fill=tk.X)

    def _on_mousewheel(self, event):
        """Обработка прокрутки колесом мыши"""
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def fill_main_frame(self, frame):
        """Заполнение основной вкладки"""
        main_canvas = tk.Canvas(frame)
        main_scrollbar = ttk.Scrollbar(frame, orient="vertical", command=main_canvas.yview)
        self.main_scrollable_frame = tk.Frame(main_canvas)

        self.main_scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=self.main_scrollable_frame, anchor="nw", width=620)
        main_canvas.configure(yscrollcommand=main_scrollbar.set)

        main_canvas.pack(side="left", fill="both", expand=True)
        main_scrollbar.pack(side="right", fill="y")

        main_canvas.bind_all("<MouseWheel>", lambda e: main_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        frame = self.main_scrollable_frame

        row = 0
        entry_width = 70
        combo_width = 68
        pady_val = 3

        # 1) Тип груза
        tk.Label(frame, text="Тип груза:").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.cargo_type = ttk.Combobox(frame, values=["уголь", "вскр. порода"], state="readonly", width=combo_width)
        self.cargo_type.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 2) Абразивность груза
        tk.Label(frame, text="Абразивность груза:").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.cargo_abrasiveness = ttk.Combobox(frame, values=["неабразивный", "малоабразивный", "абразивный",
                                                              "высокоабразивный"],
                                               state="readonly", width=combo_width)
        self.cargo_abrasiveness.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 3) Перевозка людей
        tk.Label(frame, text="Перевозка людей:").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.transport_people = ttk.Combobox(frame, values=["да", "нет"], state="readonly", width=combo_width)
        self.transport_people.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 4) Насыпная плотность груза
        tk.Label(frame, text="Насыпная плотность груза (т/м³):").grid(row=row, column=0, sticky="w", padx=5,
                                                                      pady=pady_val)
        self.density = tk.Entry(frame, width=entry_width)
        self.density.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 5) Максимальный размер куска груза
        tk.Label(frame, text="Максимальный размер куска груза (мм):").grid(row=row, column=0, sticky="w", padx=5,
                                                                           pady=pady_val)
        self.max_piece_size = tk.Entry(frame, width=entry_width)
        self.max_piece_size.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 6) Угол естественного откоса
        tk.Label(frame, text="Угол естественного откоса (градусы):").grid(row=row, column=0, sticky="w", padx=5,
                                                                          pady=pady_val)
        self.angle_of_repose = tk.Entry(frame, width=entry_width)
        self.angle_of_repose.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 7) Геометрические параметры конвейера
        tk.Label(frame, text="Геометрические параметры конвейера:",
                 font=("Arial", 9, "bold")).grid(row=row, column=0, columnspan=2, sticky="w", padx=5, pady=8)
        row += 1

        # Выбор способа ввода геометрии
        tk.Label(frame, text="Выберите способ ввода:").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.geometry_method = tk.StringVar(value="hypotenuse")
        tk.Radiobutton(frame, text="Длина и угол", variable=self.geometry_method,
                       value="hypotenuse").grid(row=row, column=1, sticky="w", padx=5, pady=pady_val)
        tk.Radiobutton(frame, text="Перепад высоты и расстояние", variable=self.geometry_method,
                       value="catheti").grid(row=row + 1, column=1, sticky="w", padx=5, pady=pady_val)
        row += 2

        # Поля для длины и угла
        self.length_label = tk.Label(frame, text="Длина конвейера (м):")
        self.length_label.grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.length_entry = tk.Entry(frame, width=entry_width)
        self.length_entry.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        self.angle_label = tk.Label(frame, text="Угол наклона конвейера (градусы):")
        self.angle_label.grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.angle_entry = tk.Entry(frame, width=entry_width)
        self.angle_entry.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # Поля для перепада высоты и расстояния
        self.height_label = tk.Label(frame, text="Перепад высоты (м):")
        self.height_label.grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.height_entry = tk.Entry(frame, width=entry_width)
        self.height_entry.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        self.height_entry.config(state="disabled")
        row += 1

        self.distance_label = tk.Label(frame, text="Расстояние (м):")
        self.distance_label.grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.distance_entry = tk.Entry(frame, width=entry_width)
        self.distance_entry.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        self.distance_entry.config(state="disabled")
        row += 1

        self.geometry_method.trace_add("write", self.update_geometry_fields)

        # 8) Потребная производительность конвейера
        tk.Label(frame, text="Потребная производительность конвейера (т/ч):").grid(row=row, column=0, sticky="w",
                                                                                   padx=5, pady=pady_val)
        self.capacity = tk.Entry(frame, width=entry_width)
        self.capacity.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 9) Скорость движущейся ленты
        tk.Label(frame, text="Скорость движущейся ленты (м/с):").grid(row=row, column=0, sticky="w", padx=5,
                                                                      pady=pady_val)
        self.speed = ttk.Combobox(frame, values=["0.5", "1.0", "1.25", "1.6", "2.0", "2.5", "3.15", "4.0"],
                                  state="readonly", width=combo_width)
        self.speed.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 10) Ширина ленты
        tk.Label(frame, text="Ширина ленты (мм):").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.belt_width = ttk.Combobox(frame, values=["500", "650", "800", "1000", "1200", "1400", "1600", "2000"],
                                       state="readonly", width=combo_width)
        self.belt_width.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 11) Тип ленты
        tk.Label(frame, text="Тип ленты:").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.belt_type = ttk.Combobox(frame, values=["резинотканевая", "резинотросовая"],
                                      state="readonly", width=combo_width)
        self.belt_type.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 12) Масса 1м² ленты
        tk.Label(frame, text="Масса 1м² ленты (кг):").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.belt_mass_per_square = tk.Entry(frame, width=entry_width)
        self.belt_mass_per_square.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 13) Минимальная температура
        tk.Label(frame, text="Минимальная температура окружающей среды:").grid(row=row, column=0, sticky="w", padx=5,
                                                                               pady=pady_val)
        self.min_temp = ttk.Combobox(frame, values=["свыше 0 градусов Цельсия", "ниже 0 градусов Цельсия"],
                                     state="readonly", width=combo_width)
        self.min_temp.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 14) Атмосферные осадки
        tk.Label(frame, text="Атмосферные осадки:").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.precipitation = ttk.Combobox(frame, values=["осадков нет", "легкие осадки", "обильные осадки"],
                                          state="readonly", width=combo_width)
        self.precipitation.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        # 15) Условия технического обслуживания
        tk.Label(frame, text="Условия технического обслуживания:").grid(row=row, column=0, sticky="w", padx=5,
                                                                        pady=pady_val)
        self.maintenance = ttk.Combobox(frame, values=["хорошее", "затруднительное"],
                                        state="readonly", width=combo_width)
        self.maintenance.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

       
        frame.columnconfigure(1, weight=1)

        tk.Label(frame, text="").grid(row=row, column=0, columnspan=2, pady=5)

    def update_geometry_fields(self, *args):
        """Обновление полей ввода геометрии при изменении метода"""
        method = self.geometry_method.get()

        if method == "hypotenuse":
            self.length_entry.config(state="normal")
            self.angle_entry.config(state="normal")
            self.height_entry.config(state="disabled")
            self.distance_entry.config(state="disabled")
            self.height_entry.delete(0, tk.END)
            self.distance_entry.delete(0, tk.END)
        else:
            self.height_entry.config(state="normal")
            self.distance_entry.config(state="normal")
            self.length_entry.config(state="disabled")
            self.angle_entry.config(state="disabled")
            self.length_entry.delete(0, tk.END)
            self.angle_entry.delete(0, tk.END)

    def fill_settings_frame(self, frame):
        """Заполнение вкладки настроек"""
        row = 0
        entry_width = 70
        pady_val = 5

        tk.Label(frame, text="Угол наклона бокового ролика (градусы):").grid(row=row, column=0, sticky="w", padx=5,
                                                                             pady=pady_val)
        self.side_roll_angle = tk.Entry(frame, width=entry_width)
        self.side_roll_angle.insert(0, str(self.settings["угол_наклона_бокового_ролика"]))
        self.side_roll_angle.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        tk.Label(frame, text="Расстояние между роликом и опорой (м):").grid(row=row, column=0, sticky="w", padx=5,
                                                                            pady=pady_val)
        self.roll_support_distance = tk.Entry(frame, width=entry_width)
        self.roll_support_distance.insert(0, str(self.settings["расстояние_между_роликом_и_опорой"]))
        self.roll_support_distance.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        tk.Label(frame, text="Количество роликов:").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.roller_count = tk.Entry(frame, width=entry_width)
        self.roller_count.insert(0, str(self.settings["количество_роликов"]))
        self.roller_count.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        tk.Label(frame, text="Масса одного ролика (кг):").grid(row=row, column=0, sticky="w", padx=5, pady=pady_val)
        self.roller_weight = tk.Entry(frame, width=entry_width)
        self.roller_weight.insert(0, str(self.settings["масса_одного_ролика"]))
        self.roller_weight.grid(row=row, column=1, padx=5, pady=pady_val, sticky="ew")
        row += 1

        save_settings_btn = tk.Button(frame, text="Сохранить настройки", command=self.save_settings_from_ui,
                                      bg="#2196F3", fg="white", font=("Arial", 10, "bold"),
                                      padx=15, pady=5)
        save_settings_btn.grid(row=row, column=0, columnspan=2, pady=10)

        frame.columnconfigure(1, weight=1)

    def save_settings_from_ui(self):
        """Сохранение настроек из UI"""
        try:
            self.settings["угол_наклона_бокового_ролика"] = float(self.side_roll_angle.get())
            self.settings["расстояние_между_роликом_и_опорой"] = float(self.roll_support_distance.get())
            self.settings["количество_роликов"] = int(self.roller_count.get())
            self.settings["масса_одного_ролика"] = float(self.roller_weight.get())

            self.save_settings()
            messagebox.showinfo("Успех", "Настройки успешно сохранены!")
        except ValueError:
            messagebox.showerror("Ошибка", "Проверьте правильность введенных данных в настройках")

    def fill_history_frame(self, frame):
        """Заполнение вкладки истории"""
        columns = ("№", "Дата", "Тип груза", "Производительность", "Мощность")
        self.history_tree = ttk.Treeview(frame, columns=columns, show="headings", height=8)

        self.history_tree.heading("№", text="№ расчёта")
        self.history_tree.heading("Дата", text="Дата и время")
        self.history_tree.heading("Тип груза", text="Тип груза")
        self.history_tree.heading("Производительность", text="Производительность (т/ч)")
        self.history_tree.heading("Мощность", text="Мощность (кВт)")

        self.history_tree.column("№", width=50)
        self.history_tree.column("Дата", width=120)
        self.history_tree.column("Тип груза", width=80)
        self.history_tree.column("Производительность", width=100)
        self.history_tree.column("Мощность", width=80)

        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)

        self.history_tree.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=5)

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        btn_frame = tk.Frame(frame)
        btn_frame.grid(row=1, column=0, columnspan=2, pady=5)

        view_btn = tk.Button(btn_frame, text="Просмотреть расчёт", command=self.view_history_item,
                             padx=8, pady=3)
        view_btn.pack(side=tk.LEFT, padx=2)

        delete_btn = tk.Button(btn_frame, text="Удалить выбранный", command=self.delete_history_item,
                               padx=8, pady=3)
        delete_btn.pack(side=tk.LEFT, padx=2)

        if EXCEL_AVAILABLE:
            export_single_excel_btn = tk.Button(btn_frame, text="Экспорт в Excel", command=self.export_single_to_excel,
                                                bg="#217346", fg="white", padx=8, pady=3)
            export_single_excel_btn.pack(side=tk.LEFT, padx=2)

        if PDF_AVAILABLE:
            export_single_pdf_btn = tk.Button(btn_frame, text="Экспорт в PDF", command=self.export_single_to_pdf,
                                              bg="#CC0000", fg="white", padx=8, pady=3)
            export_single_pdf_btn.pack(side=tk.LEFT, padx=2)

        self.refresh_history()

    def refresh_history(self):
        """Обновление таблицы истории"""
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)

        self.load_history()

        for record in self.history:
            self.history_tree.insert("", tk.END, values=(
                record.get("calculation_number", "N/A"),
                record.get("timestamp", "Неизвестно"),
                record.get("input_data", {}).get("cargo_type", "Не указано"),
                record.get("results", {}).get("capacity", "0"),
                record.get("results", {}).get("drive_power", "0")
            ))

    def view_history_item(self):
        """Просмотр выбранного элемента истории"""
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите расчёт для просмотра")
            return

        item = self.history_tree.item(selected[0])
        calc_number = item["values"][0]

        record = None
        for r in self.history:
            if r.get("calculation_number") == calc_number:
                record = r
                break

        if record:
            self.show_history_results(record)
        else:
            messagebox.showerror("Ошибка", "Запись не найдена")

    def show_history_results(self, record):
        """Отображение результатов из истории"""
        result_window = tk.Toplevel(self.root)

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = min(600, int(screen_width * 0.60))
        window_height = min(500, int(screen_height * 0.60))
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        result_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        result_window.minsize(500, 350)

        result_window.title(
            f"Расчёт №{record.get('calculation_number', 'N/A')} от {record.get('timestamp', 'Неизвестно')}")

        text_widget = tk.Text(result_window, wrap=tk.WORD, width=60, height=20)
        scrollbar = tk.Scrollbar(result_window, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)

        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        text_widget.insert(tk.END, "=" * 60 + "\n")
        text_widget.insert(tk.END, "    РЕЗУЛЬТАТЫ ТЯГОВОГО РАСЧЕТА ЛЕНТОЧНОГО КОНВЕЙЕРА\n")
        text_widget.insert(tk.END, f"    Номер расчёта: №{record.get('calculation_number', 'N/A')}\n")
        text_widget.insert(tk.END, f"    Дата расчёта: {record.get('timestamp', 'Неизвестно')}\n")
        text_widget.insert(tk.END, "=" * 60 + "\n\n")

        results = record.get("results", {})

        text_widget.insert(tk.END, "1) УСЛОВИЯ ЭКСПЛУАТАЦИИ КОНВЕЙЕРА\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        text_widget.insert(tk.END, f"   {results.get('operating_conditions', 'Нет данных')}\n\n")

        text_widget.insert(tk.END, "2) ГЕОМЕТРИЧЕСКИЕ ХАРАКТЕРИСТИКИ КОНВЕЙЕРА\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        geometry = results.get("geometry", {})
        geo_text = (
            f"   Длина конвейера (гипотенуза): {geometry.get('length', 0):.2f} м\n"
            f"   Угол наклона: {geometry.get('angle', 0):.2f}°\n"
            f"   Перепад высоты: {geometry.get('height', 0):.2f} м\n"
            f"   Горизонтальное расстояние: {geometry.get('distance', 0):.2f} м"
        )
        text_widget.insert(tk.END, f"{geo_text}\n\n")

        text_widget.insert(tk.END, "3) МИНИМАЛЬНАЯ ДОПУСТИМАЯ ШИРИНА ЛЕНТЫ\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        text_widget.insert(tk.END, f"   {results.get('min_belt_width', 0):.3f} м\n\n")

        text_widget.insert(tk.END, "4) РАСЧЕТНАЯ ПРОИЗВОДИТЕЛЬНОСТЬ\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        text_widget.insert(tk.END, f"   {results.get('capacity', 0)} т/ч\n\n")

        text_widget.insert(tk.END, "5) ПЛОЩАДЬ ПОПЕРЕЧНОГО СЕЧЕНИЯ ГРУЗА\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        text_widget.insert(tk.END, f"   {results.get('cross_section_area', 0)} м²\n\n")

        text_widget.insert(tk.END, "6) ОПРЕДЕЛЕНИЕ ПОГОДНЫХ НАГРУЗОК\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        text_widget.insert(tk.END, f"   {results.get('weather_loads', 'Нет данных')}\n\n")

        text_widget.insert(tk.END, "7) ОПРЕДЕЛЕНИЕ РАСЧЁТНОГО СОПРОТИВЛЕНИЯ ДВИЖЕНИЮ (Р)\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        text_widget.insert(tk.END, f"   {results.get('resistance_force', 0)} Н\n\n")

        text_widget.insert(tk.END, "8) РАСЧЁТНАЯ МОЩНОСТЬ ПРИВОДА\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        text_widget.insert(tk.END, f"   {results.get('drive_power', 0)} кВт\n\n")

        text_widget.insert(tk.END, "=" * 60 + "\n")

        text_widget.config(state=tk.DISABLED)

    def delete_history_item(self):
        """Удаление выбранного элемента истории"""
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите расчёт для удаления")
            return

        item = self.history_tree.item(selected[0])
        calc_number = item["values"][0]

        confirm = messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить расчёт №{calc_number}?")
        if confirm:
            self.history = [r for r in self.history if r.get("calculation_number") != calc_number]
            self.save_history()
            self.refresh_history()
            messagebox.showinfo("Успех", f"Расчёт №{calc_number} удалён из истории")

    def clear_history(self):
        """Очистка всей истории"""
        confirm = messagebox.askyesno("Подтверждение",
                                      "Вы уверены, что хотите очистить всю историю расчётов?\n\nВНИМАНИЕ: Счётчик расчётов не будет сброшен!")
        if confirm:
            self.history = []
            self.save_history()
            self.refresh_history()
            messagebox.showinfo("Успех", "История расчётов очищена")

    def calculate(self):
        """Функция выполнения расчета"""
        try:
            calc_number = self.increment_calculation_counter()
            self.counter_label.config(text=f"Следующий расчёт №: {self.calculation_number + 1}")

            data = self.collect_data()

            required_fields = [
                data['cargo_type'], data['cargo_abrasiveness'], data['transport_people'], data['density'],
                data['max_piece_size'], data['angle_of_repose'], data['capacity'],
                data['belt_type'], data['belt_mass_per_square']
            ]

            for field in required_fields:
                if not field:
                    raise ValueError("Не все обязательные поля заполнены")

            results = self.perform_calculations(data)
            self.show_results(results, calc_number, data)

            self.add_to_history(data, results, calc_number)
            self.refresh_history()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при расчете:\n{str(e)}")

    def collect_data(self):
        """Сбор данных из интерфейса"""
        data = {}

        data['cargo_type'] = self.cargo_type.get()
        data['cargo_abrasiveness'] = self.cargo_abrasiveness.get()
        data['transport_people'] = self.transport_people.get()
        data['density'] = float(self.density.get()) if self.density.get() else 0
        data['max_piece_size'] = float(self.max_piece_size.get()) if self.max_piece_size.get() else 0
        data['angle_of_repose'] = float(self.angle_of_repose.get()) if self.angle_of_repose.get() else 0
        data['capacity'] = float(self.capacity.get()) if self.capacity.get() else 0
        data['speed'] = float(self.speed.get()) if self.speed.get() else 0
        data['belt_width'] = int(self.belt_width.get()) if self.belt_width.get() else 0
        data['belt_type'] = self.belt_type.get()
        data['belt_mass_per_square'] = float(self.belt_mass_per_square.get()) if self.belt_mass_per_square.get() else 0
        data['min_temp'] = self.min_temp.get()
        data['precipitation'] = self.precipitation.get()
        data['maintenance'] = self.maintenance.get()

        method = self.geometry_method.get()
        if method == "hypotenuse":
            data['length'] = float(self.length_entry.get()) if self.length_entry.get() else 0
            data['angle'] = float(self.angle_entry.get()) if self.angle_entry.get() else 0
            angle_rad = radians(data['angle'])
            data['height'] = data['length'] * sin(angle_rad)
            data['distance'] = data['length'] * cos(angle_rad)
        else:
            data['height'] = float(self.height_entry.get()) if self.height_entry.get() else 0
            data['distance'] = float(self.distance_entry.get()) if self.distance_entry.get() else 0
            data['length'] = sqrt(data['height'] ** 2 + data['distance'] ** 2)
            data['angle'] = degrees(atan(data['height'] / data['distance'])) if data['distance'] != 0 else 0

        data['side_roll_angle'] = self.settings["угол_наклона_бокового_ролика"]
        data['roll_support_distance'] = self.settings["расстояние_между_роликом_и_опорой"]
        data['roller_count'] = self.settings["количество_роликов"]
        data['roller_weight'] = self.settings["масса_одного_ролика"]

        return data

    def perform_calculations(self, data):
        """Выполнение тяговых расчетов"""
        results = {}

        g = 9.81
        length = data.get('length', 0)
        angle_rad = radians(data.get('angle', 0))

        density = data.get('density', 0)
        belt_width = data.get('belt_width', 0) / 1000
        cross_section_area = float(self.get_cross_section_area(data))
        mass_cargo_per_meter = cross_section_area * density * 1000

        belt_mass_per_square = data.get('belt_mass_per_square', 0)
        mass_belt_per_meter = belt_mass_per_square * belt_width

        maintenance = data.get('maintenance', 'хорошее')
        if maintenance == 'хорошее':
            f = 0.02
        else:
            f = 0.035

        total_mass_per_meter = mass_cargo_per_meter + mass_belt_per_meter
        friction_resistance = f * total_mass_per_meter * g * cos(angle_rad) * length

        height = data.get('height', 0)
        lift_resistance = total_mass_per_meter * g * height

        total_resistance = friction_resistance + lift_resistance

        results['total_resistance'] = total_resistance
        results['resistance_force'] = f"{total_resistance:.2f}"

        speed = data.get('speed', 0)
        belt_type = data.get('belt_type', 'резинотканевая')
        if belt_type == 'резинотросовая':
            eta = 0.90
        else:
            eta = 0.85

        power = (total_resistance * speed) / (1000 * eta)
        power_with_reserve = power * 1.15

        results['required_power'] = power
        results['drive_power'] = f"{power_with_reserve:.2f}"

        return results

    def show_results(self, results, calc_number, input_data):
        """Отображение результатов расчета"""
        result_window = tk.Toplevel(self.root)

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = min(600, int(screen_width * 0.60))
        window_height = min(500, int(screen_height * 0.60))
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        result_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        result_window.minsize(500, 350)

        result_window.title(f"Расчёт №{calc_number}")

        btn_frame = tk.Frame(result_window)
        btn_frame.pack(pady=5)

        if EXCEL_AVAILABLE:
            export_excel_btn = tk.Button(btn_frame, text="Экспорт в Excel",
                                         command=lambda: self.export_current_to_excel(results, calc_number),
                                         bg="#217346", fg="white", padx=10, pady=3)
            export_excel_btn.pack(side=tk.LEFT, padx=5)

        if PDF_AVAILABLE:
            export_pdf_btn = tk.Button(btn_frame, text="Экспорт в PDF",
                                       command=lambda: self.export_current_to_pdf(results, calc_number),
                                       bg="#CC0000", fg="white", padx=10, pady=3)
            export_pdf_btn.pack(side=tk.LEFT, padx=5)

        text_widget = tk.Text(result_window, wrap=tk.WORD, width=60, height=20)
        scrollbar = tk.Scrollbar(result_window, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)

        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        text_widget.insert(tk.END, "=" * 60 + "\n")
        text_widget.insert(tk.END, "    РЕЗУЛЬТАТЫ ТЯГОВОГО РАСЧЕТА ЛЕНТОЧНОГО КОНВЕЙЕРА\n")
        text_widget.insert(tk.END, f"    Номер расчёта: №{calc_number}\n")
        text_widget.insert(tk.END, "=" * 60 + "\n\n")

        text_widget.insert(tk.END, "1) УСЛОВИЯ ЭКСПЛУАТАЦИИ КОНВЕЙЕРА\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        operating_conditions = self.get_operating_conditions(input_data)
        text_widget.insert(tk.END, f"   {operating_conditions}\n\n")

        text_widget.insert(tk.END, "2) ГЕОМЕТРИЧЕСКИЕ ХАРАКТЕРИСТИКИ КОНВЕЙЕРА\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        geometry_data = self.get_geometry_characteristics(input_data)
        text_widget.insert(tk.END, f"{geometry_data}\n\n")

        text_widget.insert(tk.END, "3) МИНИМАЛЬНАЯ ДОПУСТИМАЯ ШИРИНА ЛЕНТЫ\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        min_belt_width = self.get_min_belt_width(input_data)
        text_widget.insert(tk.END, f"   {min_belt_width:.3f} м\n\n")

        text_widget.insert(tk.END, "4) РАСЧЕТНАЯ ПРОИЗВОДИТЕЛЬНОСТЬ\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        calculated_capacity = self.get_calculated_capacity(input_data)
        text_widget.insert(tk.END, f"   {calculated_capacity} т/ч\n\n")

        text_widget.insert(tk.END, "5) ПЛОЩАДЬ ПОПЕРЕЧНОГО СЕЧЕНИЯ ГРУЗА\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        cross_section_area = self.get_cross_section_area(input_data)
        text_widget.insert(tk.END, f"   {cross_section_area} м²\n\n")

        text_widget.insert(tk.END, "6) ОПРЕДЕЛЕНИЕ ПОГОДНЫХ НАГРУЗОК\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        weather_loads = self.get_weather_loads(input_data)
        text_widget.insert(tk.END, f"   {weather_loads}\n\n")

        text_widget.insert(tk.END, "7) ОПРЕДЕЛЕНИЕ РАСЧЁТНОГО СОПРОТИВЛЕНИЯ ДВИЖЕНИЮ (Р)\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        resistance_force = self.get_resistance_force(results)
        text_widget.insert(tk.END, f"   {resistance_force} Н\n\n")

        text_widget.insert(tk.END, "8) РАСЧЁТНАЯ МОЩНОСТЬ ПРИВОДА\n")
        text_widget.insert(tk.END, "-" * 60 + "\n")
        drive_power = self.get_drive_power(results)
        text_widget.insert(tk.END, f"   {drive_power} кВт\n\n")

        text_widget.insert(tk.END, "=" * 60 + "\n")

        text_widget.config(state=tk.DISABLED)

    def export_single_to_excel(self):
        """Экспорт выбранного расчёта в Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Ошибка", "Библиотека openpyxl не установлена.\nУстановите: pip install openpyxl")
            return

        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите расчёт для экспорта")
            return

        item = self.history_tree.item(selected[0])
        calc_number = item["values"][0]

        record = None
        for r in self.history:
            if r.get("calculation_number") == calc_number:
                record = r
                break

        if not record:
            messagebox.showerror("Ошибка", "Запись не найдена")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"conveyor_calc_{calc_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = f"Расчёт №{calc_number}"

                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                ws.merge_cells('A1:B1')
                title_cell = ws.cell(row=1, column=1, value=f"РЕЗУЛЬТАТЫ РАСЧЁТА №{calc_number}")
                title_cell.font = Font(bold=True, size=16, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
                title_cell.alignment = center_alignment

                ws.cell(row=2, column=1, value="Дата расчёта:").font = Font(bold=True)
                ws.cell(row=2, column=2, value=record['timestamp'])

                row_num = 4

                ws.merge_cells(f'A{row_num}:B{row_num}')
                header_cell = ws.cell(row=row_num, column=1, value="ВХОДНЫЕ ДАННЫЕ")
                header_cell.font = Font(bold=True, size=12, color="FFFFFF")
                header_cell.fill = header_fill
                header_cell.alignment = center_alignment
                row_num += 1

                input_data = record.get('input_data', {})
                input_params = [
                    ["1. Тип груза", input_data.get('cargo_type', '')],
                    ["2. Абразивность груза", input_data.get('cargo_abrasiveness', '')],
                    ["3. Перевозка людей", input_data.get('transport_people', '')],
                    ["4. Насыпная плотность (т/м³)", input_data.get('density', '')],
                    ["5. Макс. размер куска (мм)", input_data.get('max_piece_size', '')],
                    ["6. Угол естественного откоса (°)", input_data.get('angle_of_repose', '')],
                    ["7. Длина конвейера (м)", f"{input_data.get('length', 0):.2f}"],
                    ["8. Угол наклона (°)", f"{input_data.get('angle', 0):.2f}"],
                    ["9. Перепад высоты (м)", f"{input_data.get('height', 0):.2f}"],
                    ["10. Горизонтальное расстояние (м)", f"{input_data.get('distance', 0):.2f}"],
                    ["11. Потребная производительность (т/ч)", input_data.get('capacity', '')],
                    ["12. Скорость ленты (м/с)", input_data.get('speed', '')],
                    ["13. Ширина ленты (мм)", input_data.get('belt_width', '')],
                    ["14. Тип ленты", input_data.get('belt_type', '')],
                    ["15. Масса 1м² ленты (кг)", input_data.get('belt_mass_per_square', '')],
                    ["16. Мин. температура", input_data.get('min_temp', '')],
                    ["17. Атмосферные осадки", input_data.get('precipitation', '')],
                    ["18. Условия обслуживания", input_data.get('maintenance', '')],
                ]

                for param, value in input_params:
                    ws.cell(row=row_num, column=1, value=param).font = Font(bold=True)
                    ws.cell(row=row_num, column=2, value=value)
                    row_num += 1

                row_num += 1

                ws.merge_cells(f'A{row_num}:B{row_num}')
                header_cell = ws.cell(row=row_num, column=1, value="РЕЗУЛЬТАТЫ РАСЧЁТА")
                header_cell.font = Font(bold=True, size=12, color="FFFFFF")
                header_cell.fill = header_fill
                header_cell.alignment = center_alignment
                row_num += 1

                results = record.get('results', {})
                geometry = results.get('geometry', {})

                result_params = [
                    ["1. Условия эксплуатации", results.get('operating_conditions', '')],
                    ["2. Длина конвейера (м)", f"{geometry.get('length', 0):.2f}"],
                    ["3. Угол наклона (°)", f"{geometry.get('angle', 0):.2f}"],
                    ["4. Перепад высоты (м)", f"{geometry.get('height', 0):.2f}"],
                    ["5. Горизонтальное расстояние (м)", f"{geometry.get('distance', 0):.2f}"],
                    ["6. Мин. ширина ленты (м)", results.get('min_belt_width', '')],
                    ["7. Расчётная производительность (т/ч)", results.get('capacity', '')],
                    ["8. Площадь поперечного сечения (м²)", results.get('cross_section_area', '')],
                    ["9. Погодные нагрузки", results.get('weather_loads', '')],
                    ["10. Сопротивление движению Р (Н)", results.get('resistance_force', '')],
                    ["11. Мощность привода (кВт)", results.get('drive_power', '')],
                ]

                for param, value in result_params:
                    ws.cell(row=row_num, column=1, value=param).font = Font(bold=True)
                    ws.cell(row=row_num, column=2, value=value)
                    row_num += 1

                for row in ws.iter_rows(min_row=1, max_row=row_num, max_col=2):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                ws.column_dimensions['A'].width = 45
                ws.column_dimensions['B'].width = 35

                wb.save(file_path)
                messagebox.showinfo("Успех", f"Расчёт №{calc_number} экспортирован в Excel")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при экспорте:\n{str(e)}")

    def export_single_to_pdf(self):
        """Экспорт выбранного расчёта в PDF"""
        if not PDF_AVAILABLE:
            messagebox.showerror("Ошибка", "Библиотека reportlab не установлена.\nУстановите: pip install reportlab")
            return

        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите расчёт для экспорта")
            return

        item = self.history_tree.item(selected[0])
        calc_number = item["values"][0]

        record = None
        for r in self.history:
            if r.get("calculation_number") == calc_number:
                record = r
                break

        if not record:
            messagebox.showerror("Ошибка", "Запись не найдена")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=f"conveyor_calc_{calc_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if file_path:
            try:
                doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                                        topMargin=2 * cm, bottomMargin=2 * cm)
                styles = getSampleStyleSheet()

                title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=darkblue,
                                             spaceAfter=20, alignment=TA_CENTER, fontName=self.cyrillic_font)
                heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=12, textColor=black,
                                               spaceAfter=10, spaceBefore=10, fontName=self.cyrillic_font)
                normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontSize=9, spaceAfter=4,
                                              fontName=self.cyrillic_font)

                elements = []

                elements.append(Paragraph(f"РЕЗУЛЬТАТЫ РАСЧЁТА №{calc_number}", title_style))
                elements.append(Paragraph(f"Дата: {record.get('timestamp', '')}", normal_style))
                elements.append(Spacer(1, 0.3 * cm))

                elements.append(Paragraph("ВХОДНЫЕ ДАННЫЕ:", heading_style))
                input_data = record.get('input_data', {})

                input_table_data = [
                    ['Параметр', 'Значение'],
                    ['1. Тип груза', str(input_data.get('cargo_type', ''))],
                    ['2. Абразивность груза', str(input_data.get('cargo_abrasiveness', ''))],
                    ['3. Перевозка людей', str(input_data.get('transport_people', ''))],
                    ['4. Насыпная плотность (т/м³)', str(input_data.get('density', ''))],
                    ['5. Макс. размер куска (мм)', str(input_data.get('max_piece_size', ''))],
                    ['6. Угол естественного откоса (°)', str(input_data.get('angle_of_repose', ''))],
                    ['7. Длина конвейера (м)', f"{input_data.get('length', 0):.2f}"],
                    ['8. Угол наклона (°)', f"{input_data.get('angle', 0):.2f}"],
                    ['9. Перепад высоты (м)', f"{input_data.get('height', 0):.2f}"],
                    ['10. Горизонтальное расстояние (м)', f"{input_data.get('distance', 0):.2f}"],
                    ['11. Производительность (т/ч)', str(input_data.get('capacity', ''))],
                    ['12. Скорость ленты (м/с)', str(input_data.get('speed', ''))],
                    ['13. Ширина ленты (мм)', str(input_data.get('belt_width', ''))],
                    ['14. Тип ленты', str(input_data.get('belt_type', ''))],
                    ['15. Масса 1м² ленты (кг)', str(input_data.get('belt_mass_per_square', ''))],
                    ['16. Мин. температура', str(input_data.get('min_temp', ''))],
                    ['17. Атмосферные осадки', str(input_data.get('precipitation', ''))],
                    ['18. Условия обслуживания', str(input_data.get('maintenance', ''))],
                ]

                input_table = Table(input_table_data, colWidths=[5.5 * cm, 5.5 * cm])
                input_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), self.cyrillic_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), white),
                    ('GRID', (0, 0), (-1, -1), 0.5, gray),
                ]))
                elements.append(input_table)
                elements.append(Spacer(1, 0.3 * cm))

                elements.append(Paragraph("РЕЗУЛЬТАТЫ РАСЧЁТА:", heading_style))
                results = record.get('results', {})
                geometry = results.get('geometry', {})

                results_table_data = [
                    ['Параметр', 'Значение'],
                    ['1. Условия эксплуатации', str(results.get('operating_conditions', ''))],
                    ['2. Длина конвейера (м)', f"{geometry.get('length', 0):.2f}"],
                    ['3. Угол наклона (°)', f"{geometry.get('angle', 0):.2f}"],
                    ['4. Перепад высоты (м)', f"{geometry.get('height', 0):.2f}"],
                    ['5. Горизонтальное расстояние (м)', f"{geometry.get('distance', 0):.2f}"],
                    ['6. Мин. ширина ленты (м)', str(results.get('min_belt_width', ''))],
                    ['7. Производительность (т/ч)', str(results.get('capacity', ''))],
                    ['8. Площадь сечения (м²)', str(results.get('cross_section_area', ''))],
                    ['9. Погодные нагрузки', str(results.get('weather_loads', ''))],
                    ['10. Сопротивление движению Р (Н)', str(results.get('resistance_force', ''))],
                    ['11. Мощность привода (кВт)', str(results.get('drive_power', ''))],
                ]

                results_table = Table(results_table_data, colWidths=[5.5 * cm, 5.5 * cm])
                results_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), self.cyrillic_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), white),
                    ('GRID', (0, 0), (-1, -1), 0.5, gray),
                ]))
                elements.append(results_table)

                doc.build(elements)
                messagebox.showinfo("Успех", f"Расчёт №{calc_number} экспортирован в PDF")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при экспорте в PDF:\n{str(e)}")

    def export_current_to_excel(self, results, calc_number):
        """Экспорт текущего расчёта в Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Ошибка", "Библиотека openpyxl не установлена.")
            return

        input_data = self.collect_data()

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"conveyor_calc_{calc_number}.xlsx"
        )

        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = f"Расчёт №{calc_number}"

                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                ws.merge_cells('A1:B1')
                title_cell = ws.cell(row=1, column=1, value=f"РЕЗУЛЬТАТЫ РАСЧЁТА №{calc_number}")
                title_cell.font = Font(bold=True, size=16, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
                title_cell.alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(row=2, column=1, value="Дата расчёта:").font = Font(bold=True)
                ws.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

                row_num = 4

                ws.merge_cells(f'A{row_num}:B{row_num}')
                header_cell = ws.cell(row=row_num, column=1, value="ВХОДНЫЕ ДАННЫЕ")
                header_cell.font = Font(bold=True, size=12, color="FFFFFF")
                header_cell.fill = header_fill
                header_cell.alignment = Alignment(horizontal='center')
                row_num += 1

                input_params = [
                    ["1. Тип груза", input_data.get('cargo_type', '')],
                    ["2. Абразивность груза", input_data.get('cargo_abrasiveness', '')],
                    ["3. Перевозка людей", input_data.get('transport_people', '')],
                    ["4. Насыпная плотность (т/м³)", input_data.get('density', '')],
                    ["5. Макс. размер куска (мм)", input_data.get('max_piece_size', '')],
                    ["6. Угол естественного откоса (°)", input_data.get('angle_of_repose', '')],
                    ["7. Длина конвейера (м)", f"{input_data.get('length', 0):.2f}"],
                    ["8. Угол наклона (°)", f"{input_data.get('angle', 0):.2f}"],
                    ["9. Перепад высоты (м)", f"{input_data.get('height', 0):.2f}"],
                    ["10. Горизонтальное расстояние (м)", f"{input_data.get('distance', 0):.2f}"],
                    ["11. Производительность (т/ч)", input_data.get('capacity', '')],
                    ["12. Скорость ленты (м/с)", input_data.get('speed', '')],
                    ["13. Ширина ленты (мм)", input_data.get('belt_width', '')],
                    ["14. Тип ленты", input_data.get('belt_type', '')],
                    ["15. Масса 1м² ленты (кг)", input_data.get('belt_mass_per_square', '')],
                    ["16. Мин. температура", input_data.get('min_temp', '')],
                    ["17. Атмосферные осадки", input_data.get('precipitation', '')],
                    ["18. Условия обслуживания", input_data.get('maintenance', '')],
                ]

                for param, value in input_params:
                    ws.cell(row=row_num, column=1, value=param).font = Font(bold=True)
                    ws.cell(row=row_num, column=2, value=value)
                    row_num += 1

                row_num += 1

                ws.merge_cells(f'A{row_num}:B{row_num}')
                header_cell = ws.cell(row=row_num, column=1, value="РЕЗУЛЬТАТЫ РАСЧЁТА")
                header_cell.font = Font(bold=True, size=12, color="FFFFFF")
                header_cell.fill = header_fill
                header_cell.alignment = Alignment(horizontal='center')
                row_num += 1

                result_params = [
                    ["1. Условия эксплуатации", self.get_operating_conditions(input_data)],
                    ["2. Длина конвейера (м)", f"{input_data.get('length', 0):.2f}"],
                    ["3. Угол наклона (°)", f"{input_data.get('angle', 0):.2f}"],
                    ["4. Перепад высоты (м)", f"{input_data.get('height', 0):.2f}"],
                    ["5. Горизонтальное расстояние (м)", f"{input_data.get('distance', 0):.2f}"],
                    ["6. Мин. ширина ленты (м)", self.get_min_belt_width(input_data)],
                    ["7. Производительность (т/ч)", self.get_calculated_capacity(input_data)],
                    ["8. Площадь сечения (м²)", self.get_cross_section_area(input_data)],
                    ["9. Погодные нагрузки", self.get_weather_loads(input_data)],
                    ["10. Сопротивление движению Р (Н)", self.get_resistance_force(results)],
                    ["11. Мощность привода (кВт)", self.get_drive_power(results)],
                ]

                for param, value in result_params:
                    ws.cell(row=row_num, column=1, value=param).font = Font(bold=True)
                    ws.cell(row=row_num, column=2, value=value)
                    row_num += 1

                for row in ws.iter_rows(min_row=1, max_row=row_num, max_col=2):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                ws.column_dimensions['A'].width = 45
                ws.column_dimensions['B'].width = 35

                wb.save(file_path)
                messagebox.showinfo("Успех", f"Расчёт экспортирован в Excel")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при экспорте:\n{str(e)}")

    def export_current_to_pdf(self, results, calc_number):
        """Экспорт текущего расчёта в PDF"""
        if not PDF_AVAILABLE:
            messagebox.showerror("Ошибка", "Библиотека reportlab не установлена.")
            return

        input_data = self.collect_data()

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"conveyor_calc_{calc_number}.pdf"
        )

        if file_path:
            try:
                doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                                        topMargin=2 * cm, bottomMargin=2 * cm)
                styles = getSampleStyleSheet()

                title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=darkblue,
                                             spaceAfter=20, alignment=TA_CENTER, fontName=self.cyrillic_font)
                heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=12, textColor=black,
                                               spaceAfter=10, fontName=self.cyrillic_font)
                normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontSize=9, spaceAfter=4,
                                              fontName=self.cyrillic_font)

                elements = []

                elements.append(Paragraph(f"РЕЗУЛЬТАТЫ РАСЧЁТА №{calc_number}", title_style))
                elements.append(Paragraph(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
                elements.append(Spacer(1, 0.3 * cm))

                elements.append(Paragraph("ВХОДНЫЕ ДАННЫЕ:", heading_style))
                input_table_data = [
                    ['Параметр', 'Значение'],
                    ['1. Тип груза', str(input_data.get('cargo_type', ''))],
                    ['2. Абразивность груза', str(input_data.get('cargo_abrasiveness', ''))],
                    ['3. Перевозка людей', str(input_data.get('transport_people', ''))],
                    ['4. Насыпная плотность (т/м³)', str(input_data.get('density', ''))],
                    ['5. Макс. размер куска (мм)', str(input_data.get('max_piece_size', ''))],
                    ['6. Угол естественного откоса (°)', str(input_data.get('angle_of_repose', ''))],
                    ['7. Длина конвейера (м)', f"{input_data.get('length', 0):.2f}"],
                    ['8. Угол наклона (°)', f"{input_data.get('angle', 0):.2f}"],
                    ['9. Перепад высоты (м)', f"{input_data.get('height', 0):.2f}"],
                    ['10. Горизонтальное расстояние (м)', f"{input_data.get('distance', 0):.2f}"],
                    ['11. Производительность (т/ч)', str(input_data.get('capacity', ''))],
                    ['12. Скорость ленты (м/с)', str(input_data.get('speed', ''))],
                    ['13. Ширина ленты (мм)', str(input_data.get('belt_width', ''))],
                    ['14. Тип ленты', str(input_data.get('belt_type', ''))],
                    ['15. Масса 1м² ленты (кг)', str(input_data.get('belt_mass_per_square', ''))],
                    ['16. Мин. температура', str(input_data.get('min_temp', ''))],
                    ['17. Атмосферные осадки', str(input_data.get('precipitation', ''))],
                    ['18. Условия обслуживания', str(input_data.get('maintenance', ''))],
                ]

                input_table = Table(input_table_data, colWidths=[5.5 * cm, 5.5 * cm])
                input_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), self.cyrillic_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), white),
                    ('GRID', (0, 0), (-1, -1), 0.5, gray),
                ]))
                elements.append(input_table)
                elements.append(Spacer(1, 0.3 * cm))

                elements.append(Paragraph("РЕЗУЛЬТАТЫ РАСЧЁТА:", heading_style))
                results_table_data = [
                    ['Параметр', 'Значение'],
                    ['1. Условия эксплуатации', str(self.get_operating_conditions(input_data))],
                    ['2. Длина конвейера (м)', f"{input_data.get('length', 0):.2f}"],
                    ['3. Угол наклона (°)', f"{input_data.get('angle', 0):.2f}"],
                    ['4. Перепад высоты (м)', f"{input_data.get('height', 0):.2f}"],
                    ['5. Горизонтальное расстояние (м)', f"{input_data.get('distance', 0):.2f}"],
                    ['6. Мин. ширина ленты (м)', str(self.get_min_belt_width(input_data))],
                    ['7. Производительность (т/ч)', str(self.get_calculated_capacity(input_data))],
                    ['8. Площадь сечения (м²)', str(self.get_cross_section_area(input_data))],
                    ['9. Погодные нагрузки', str(self.get_weather_loads(input_data))],
                    ['10. Сопротивление движению Р (Н)', self.get_resistance_force(results)],
                    ['11. Мощность привода (кВт)', self.get_drive_power(results)],
                ]

                results_table = Table(results_table_data, colWidths=[5.5 * cm, 5.5 * cm])
                results_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), self.cyrillic_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), white),
                    ('GRID', (0, 0), (-1, -1), 0.5, gray),
                ]))
                elements.append(results_table)

                doc.build(elements)
                messagebox.showinfo("Успех", f"Расчёт экспортирован в PDF")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при экспорте:\n{str(e)}")

    def get_operating_conditions(self, data):
        """1) Условия эксплуатации конвейера"""
        max_piece = data.get('max_piece_size', 0)
        abraz = data.get('cargo_abrasiveness', '')
        density = data.get('density', 0)
        min_temp = data.get('min_temp', 'Не указано')
        precipitation = data.get('precipitation', 'Не указано')
        maintenance = data.get('maintenance', 'Не указано')

        conditions = ''
        res = 0
        gruz = 0
        abr = 0

        if max_piece:
            if 80 < max_piece <= 150:
                gruz = 8
            elif 150 < max_piece <= 350:
                gruz = 18
            elif 350 < max_piece <= 500:
                gruz = 25

        if abraz == 'неабразивный':
            abr = 0
        elif abraz == 'малоабразивный':
            abr = 5
        elif abraz == 'абразивный':
            abr = 15
        elif abraz == 'высокоабразивный':
            abr = 25

        if min_temp == 'ниже 0 градусов Цельсия':
            res += 10
        if precipitation == 'обильные осадки':
            res += 10
        if maintenance == 'затруднительное':
            res += 20

        if density:
            if 1 < density <= 1.7:
                gruz *= 1.7
            elif 1.7 < density <= 2.3:
                gruz *= 1.9
            elif 2.3 < density <= 2.7:
                gruz *= 2
            elif 2.7 < density:
                gruz *= 2.2

        res += (gruz + abr)

        if res <= 20:
            conditions = 'условия эксплуатации лёгкие'
        elif 20 < res <= 50:
            conditions = 'условия эксплуатации средние'
        elif 50 < res <= 75:
            conditions = 'условия эксплуатации тяжелые'
        elif 75 < res <= 100:
            conditions = 'условия эксплуатации очень тяжелые'
        else:
            conditions = 'применение конвейерной ленты не допускается'

        return conditions

    def get_geometry_characteristics(self, data):
        """2) Геометрические характеристики конвейера"""
        length = data.get('length', 0)
        angle = data.get('angle', 0)
        height = data.get('height', 0)
        distance = data.get('distance', 0)

        geometry = (
            f"   Длина конвейера (гипотенуза): {length:.2f} м\n"
            f"   Угол наклона конвейера: {angle:.2f}°\n"
            f"   Перепад высоты: {height:.2f} м\n"
            f"   Горизонтальное расстояние: {distance:.2f} м"
        )

        return geometry

    def get_min_belt_width(self, data):
        """3) Минимальная допустимая ширина ленты"""
        capacity = data.get('caapcity', 0)
        speed = data.get('speed', 0)
        density = data.get('density', 0)

        min_width = 1.1*(math.sqrt(((capacity * 1.2)/(0.9*0.96))/(speed*density*660*0.95))+0.05)
        return min_width

    def get_calculated_capacity(self, data):
        """4) Расчетная производительность"""

        capacity = data.get('capacity', 0)
        res = (capacity * 1.2)/(0.9*0.96)
        return res

    def calculate_cross_section_area(self, data):
        """5) Расчёт площади поперечного сечения груза (по формуле из Excel)

        F = (b²/4) * {[cos β' + θ(1 - cos β')]² * (tan φ₀ + tan β') - θ² * tan β'}
        где θ = lр / b
        """
        B = data.get('belt_width', 1000) / 1000

        phi_0_deg = data.get('angle_of_repose', 15)
        beta_prime_deg = data.get('side_roll_angle', 35)
        l_p = data.get('roller_length', 0.36)

        b = B - 0.15

        theta = l_p / b

        phi_0_rad = radians(phi_0_deg)
        beta_prime_rad = radians(beta_prime_deg)


        cos_beta = cos(beta_prime_rad)
        tan_phi_0 = tan(phi_0_rad)
        tan_beta = tan(beta_prime_rad)
        # F = (b²/4) * {[cos β' + θ(1 - cos β')]² * (tan φ₀ + tan β') - θ² * tan β'}

        term1 = cos_beta + theta * (1 - cos_beta)
        term1_squared = term1 ** 2
        term2 = tan_phi_0 + tan_beta
        term3 = theta ** 2 * tan_beta
        F = (b ** 2 / 4) * (term1_squared * term2 - term3)

        return f"{max(0, F):.3f}"

    def get_weather_loads(self, data):
        """6) Определение погодных нагрузок"""
        min_temp = data.get('min_temp', 'Не указано')
        precipitation = data.get('precipitation', 'Не указано')
        weather_info = f"Температурный режим: {min_temp} | Осадки: {precipitation}"
        return weather_info

    def get_resistance_coefficient(self, data, operating_conditions):
        """Определение коэффициента сопротивления движению ω (по таблице 14)

        Зависит от:
        - Условий эксплуатации (лёгкие, средние, тяжёлые, очень тяжёлые)
        - Сезона (зима/лето)
        - Длины конвейера (до 100 м или свыше)
        """
        # Длина конвейера
        L = data.get('length', 0)

        # Определение сезона
        min_temp = data.get('min_temp', '')
        is_winter = 'ниже 0' in min_temp

        # Определение коэффициента ω из таблицы 14
        if 'лёгкие' in operating_conditions:
            if L <= 100:
                omega = 0.02
            else:
                omega = 0.018

        elif 'средние' in operating_conditions:
            if L <= 100:
                omega = 0.025
            else:
                omega = 0.022

        elif 'тяжёлые' in operating_conditions:
            if is_winter:
                if L <= 100:
                    omega = 0.045
                else:
                    omega = 0.042
            else:
                if L <= 100:
                    omega = 0.035
                else:
                    omega = 0.032

        elif 'очень тяжёлые' in operating_conditions:
            if is_winter:
                if L <= 100:
                    omega = 0.055
                else:
                    omega = 0.05
            else:
                if L <= 100:
                    omega = 0.04
                else:
                    omega = 0.036

        else:
            omega = 0.035

        return omega

    def calculate_resistance_force(self, data):
        """7) Расчёт сопротивления движению ленты (по формуле из Excel)

        P = gL[(qг + q'р + q''р + 2qл) · ωcosβ + qгsinβ]
        """
        g = 9.81

        L = data.get('length', 0)

        roller_count_upper = 3  # из Excel
        roller_weight = data.get('roller_weight', 9.5)
        roll_distance_upper = data.get('roll_support_distance', 1.5)
        q_r_upper = (roller_count_upper * roller_weight) / roll_distance_upper

        roller_count_lower = 2
        roll_distance_lower = roll_distance_upper * 2
        q_r_lower = (roller_count_lower * roller_weight) / roll_distance_lower
        q_l = data.get('belt_mass_per_square', 24)
        capacity = float(self.calculate_capacity(data, self.calculate_cross_section_area(data)))
        v = data.get('speed', 1.0)
        q_g = capacity / (3.6 * v) if v > 0 else 0

        beta_deg = data.get('angle', 0)
        beta_rad = radians(beta_deg)
        operating_conditions = self.get_operating_conditions(data)
        omega = self.get_resistance_coefficient(data, operating_conditions)

        # Основная формула:
        # P = gL[(qг + q'р + q''р + 2qл) · ωcosβ + qгsinβ]
        total_mass = q_g + q_r_upper + q_r_lower + 2 * q_l

        # Расчёт по частям
        friction_component = total_mass * omega * cos(beta_rad)
        lift_component = q_g * sin(beta_rad)
        P = g * L * (friction_component + lift_component) / 1000

        return f"{abs(P):.1f}"

    def calculate_drive_power(self, data, resistance_force_str):
        """8) Расчёт мощности привода (по формуле из Excel)

        Nгр = (P * v * K) / h
        где:
        - P - расчётное сопротивление движению, кН
        - v - скорость ленты, м/с
        - K = 1.32 - коэффициент запаса мощности
        - h = 0.85 - КПД двигателя
        """
        try:
            P = float(resistance_force_str)  # сопротивление в кН
        except:
            P = 0

        v = data.get('speed', 0)  # скорость ленты, м/с

        K = 1.32
        h = 0.85

        # Формула: Nгр = (P * v * K) / h
        power = (P * v * K) / h

        return f"{power:.1f}"

    def run(self):
        """Запуск приложения"""
        self.root.mainloop()


# Запуск приложения
if __name__ == "__main__":
    app = ConveyorCalculator()
    app.run()